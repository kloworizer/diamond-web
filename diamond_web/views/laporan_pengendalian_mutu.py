"""Laporan Pengendalian Mutu view - Quality Control Report."""

from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.decorators.http import require_GET, require_http_methods
from django.views.generic import TemplateView
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_protect
from datetime import datetime, timedelta
from django.db.models import Q
from io import BytesIO
from openpyxl import Workbook

from ..models.tiket import Tiket
from ..constants.tiket_status import STATUS_LABELS
from ..forms.laporan_pengendalian_mutu import LaporanPengendalianMutuFilterForm, TiketExportResource


def _is_pmde_user(user):
    """Check if user is PMDE user or admin."""
    return user.is_superuser or user.is_staff or user.groups.filter(name__in=['user_pmde', 'admin', 'admin_pmde']).exists()


class LaporanPengendalianMutuView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Display Quality Control Report with filtering by quarter (triwulan) and year.
    
    This view renders a template with filter form and DataTables table that displays
    tikets transferred to PMDE within a specified quarter and year. Access control
    is enforced via test_func() to verify user is PMDE user or admin.
    
    Template: laporan_pengendalian_mutu/list.html
    """
    template_name = 'laporan_pengendalian_mutu/list.html'
    
    def test_func(self):
        """Verify user is PMDE user or admin."""
        return _is_pmde_user(self.request.user)
    
    def get_context_data(self, **kwargs):
        """Add years list and form to context for filter dropdown."""
        context = super().get_context_data(**kwargs)
        # Get unique years from tiket data
        tikets = Tiket.objects.all()
        years = sorted(set(t.tahun for t in tikets), reverse=True)
        
        # Always include current year
        current_year = datetime.now().year
        if current_year not in years:
            years = [current_year] + list(years)
        
        context['years'] = years
        # Initialize form with years
        context['form'] = LaporanPengendalianMutuFilterForm(years=years)
        return context


@login_required
@user_passes_test(_is_pmde_user)
@require_http_methods(["GET", "POST"])
@csrf_protect
def laporan_pengendalian_mutu_data(request):
    """DataTables server-side endpoint for Quality Control Report.
    
    Filters tikets by:
    - tgl_transfer within specified periode (month/quarter/semester/year) and tahun (year)
    - Periode types:
      - bulanan (monthly): Filter by specific month (1-12)
      - triwulanan (quarterly): Filter by quarter (1-4)
      - semester (semiannual): Filter by semester (1-2)
      - tahunan (yearly): Filter by entire year
    
    GET Parameters:
    - periode_type: Type of period (bulanan, triwulanan, semester, tahunan), required
    - periode: Period value based on type, required
    - tahun: Year, required
    - draw: DataTables draw counter
    - start: Record offset for pagination
    - length: Number of records per page
    
    Returns JSON with tiket data including QC fields.
    """
    params = request.POST if request.method == 'POST' else request.GET
    periode_type = params.get('periode_type')
    periode = params.get('periode')
    tahun = params.get('tahun')

    try:
        draw = int(params.get('draw', 1))
    except (ValueError, TypeError):
        draw = 1

    try:
        start = int(params.get('start', 0))
    except (ValueError, TypeError):
        start = 0

    try:
        length = int(params.get('length', 10))
    except (ValueError, TypeError):
        length = 10
    
    # Validate inputs
    if not periode_type or not periode or not tahun:
        return JsonResponse({
            'draw': draw,
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': []
        })
    
    try:
        tahun = int(tahun)
    except (ValueError, TypeError):
        return JsonResponse({
            'draw': draw,
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': []
        })
    
    # Calculate date range based on periode type
    start_date = None
    end_date = None
    
    if periode_type == 'bulanan':
        try:
            bulan = int(periode)
            if bulan < 1 or bulan > 12:
                return JsonResponse({
                    'draw': draw,
                    'recordsTotal': 0,
                    'recordsFiltered': 0,
                    'data': []
                })
            start_date = datetime(tahun, bulan, 1).date()
            # Calculate end date for the month
            if bulan == 12:
                end_date = datetime(tahun + 1, 1, 1).date() - timedelta(days=1)
            else:
                end_date = datetime(tahun, bulan + 1, 1).date() - timedelta(days=1)
        except (ValueError, TypeError):
            return JsonResponse({
                'draw': draw,
                'recordsTotal': 0,
                'recordsFiltered': 0,
                'data': []
            })
    
    elif periode_type == 'triwulanan':
        quarter_months = {
            1: (1, 3),      # Q1: Jan-Mar
            2: (4, 6),      # Q2: Apr-Jun
            3: (7, 9),      # Q3: Jul-Sep
            4: (10, 12),    # Q4: Oct-Dec
        }
        try:
            triwulan = int(periode)
            if triwulan not in quarter_months:
                return JsonResponse({
                    'draw': draw,
                    'recordsTotal': 0,
                    'recordsFiltered': 0,
                    'data': []
                })
            start_month, end_month = quarter_months[triwulan]
            start_date = datetime(tahun, start_month, 1).date()
            if end_month == 12:
                end_date = datetime(tahun + 1, 1, 1).date() - timedelta(days=1)
            else:
                end_date = datetime(tahun, end_month + 1, 1).date() - timedelta(days=1)
        except (ValueError, TypeError):
            return JsonResponse({
                'draw': draw,
                'recordsTotal': 0,
                'recordsFiltered': 0,
                'data': []
            })
    
    elif periode_type == 'semester':
        semester_months = {
            1: (1, 6),      # Semester 1: Jan-Jun
            2: (7, 12),     # Semester 2: Jul-Dec
        }
        try:
            semester = int(periode)
            if semester not in semester_months:
                return JsonResponse({
                    'draw': draw,
                    'recordsTotal': 0,
                    'recordsFiltered': 0,
                    'data': []
                })
            start_month, end_month = semester_months[semester]
            start_date = datetime(tahun, start_month, 1).date()
            if end_month == 12:
                end_date = datetime(tahun + 1, 1, 1).date() - timedelta(days=1)
            else:
                end_date = datetime(tahun, end_month + 1, 1).date() - timedelta(days=1)
        except (ValueError, TypeError):
            return JsonResponse({
                'draw': draw,
                'recordsTotal': 0,
                'recordsFiltered': 0,
                'data': []
            })
    
    elif periode_type == 'tahunan':
        start_date = datetime(tahun, 1, 1).date()
        end_date = datetime(tahun, 12, 31).date()
    
    else:
        return JsonResponse({
            'draw': draw,
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': []
        })
    
    # Query tikets with tgl_transfer in the specified period
    tikets = Tiket.objects.filter(
        tgl_transfer__isnull=False,
        tgl_transfer__date__gte=start_date,
        tgl_transfer__date__lte=end_date
    ).select_related(
        'id_periode_data__id_sub_jenis_data_ilap__id_ilap',
        'id_periode_data__id_sub_jenis_data_ilap__id_jenis_tabel'
    ).order_by('-tgl_transfer')
    
    records_total = Tiket.objects.count()
    records_filtered = tikets.count()
    
    # Pagination
    tikets = tikets[start:start + length]
    
    # Build response data
    data = []
    for tiket in tikets:
        sub_jenis_data = tiket.id_periode_data.id_sub_jenis_data_ilap
        ilap = sub_jenis_data.id_ilap
        jenis_tabel = sub_jenis_data.id_jenis_tabel
        
        row = {
            'nama_ilap': ilap.nama_ilap,
            'nama_sub_jenis_data': sub_jenis_data.nama_sub_jenis_data,
            'nama_tabel': jenis_tabel.deskripsi if jenis_tabel else '',
            'nomor_tiket': tiket.nomor_tiket,
            'status_tiket': STATUS_LABELS.get(tiket.status_tiket, 'Unknown'),
            'data_diterima': tiket.baris_diterima or 0,
            'data_direkam': (tiket.baris_i or 0) + (tiket.baris_u or 0),
            'data_teridentifikasi_i': tiket.baris_i or 0,
            'data_tidak_teridentifikasi_u': tiket.baris_u or 0,
            'lolos_qc': tiket.lolos_qc or 0,
            'tidak_lolos_qc': tiket.tidak_lolos_qc or 0,
            'qc_p': tiket.qc_p or 0,
            'qc_x': tiket.qc_x or 0,
            'qc_w': tiket.qc_w or 0,
            'qc_v': tiket.qc_v or 0,
            'qc_a': tiket.qc_a or 0,
            'qc_n': tiket.qc_n or 0,
            'qc_y': tiket.qc_y or 0,
            'qc_z': tiket.qc_z or 0,
            'qc_d': tiket.qc_d or 0,
            'qc_u': tiket.qc_u or 0,
            'qc_c': tiket.qc_c or 0,
        }
        data.append(row)
    
    return JsonResponse({
        'draw': draw,
        'recordsTotal': records_total,
        'recordsFiltered': records_filtered,
        'data': data
    })


@login_required
@user_passes_test(_is_pmde_user)
@require_GET
@csrf_protect
def laporan_pengendalian_mutu_export(request):
    """Export Laporan Pengendalian Mutu to XLSX file.
    
    GET Parameters:
    - periode_type: Type of period (bulanan, triwulanan, semester, tahunan)
    - periode: Period value
    - tahun: Year
    
    Returns: XLSX file download
    """
    periode_type = request.GET.get('periode_type')
    periode = request.GET.get('periode')
    tahun = request.GET.get('tahun')
    
    # Validate inputs
    if not periode_type or not periode or not tahun:
        return HttpResponse('Invalid parameters', status=400)
    
    try:
        tahun = int(tahun)
    except (ValueError, TypeError):
        return HttpResponse('Invalid year', status=400)
    
    # Calculate date range based on periode type
    start_date = None
    end_date = None
    periode_label = ''
    
    if periode_type == 'bulanan':
        try:
            bulan = int(periode)
            if bulan < 1 or bulan > 12:
                return HttpResponse('Invalid month', status=400)
            start_date = datetime(tahun, bulan, 1).date()
            if bulan == 12:
                end_date = datetime(tahun + 1, 1, 1).date() - timedelta(days=1)
            else:
                end_date = datetime(tahun, bulan + 1, 1).date() - timedelta(days=1)
            
            bulan_names = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                          'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
            periode_label = f'{bulan_names[bulan]} {tahun}'
        except (ValueError, TypeError):
            return HttpResponse('Invalid month', status=400)
    
    elif periode_type == 'triwulanan':
        quarter_months = {
            1: (1, 3),
            2: (4, 6),
            3: (7, 9),
            4: (10, 12),
        }
        try:
            triwulan = int(periode)
            if triwulan not in quarter_months:
                return HttpResponse('Invalid quarter', status=400)
            start_month, end_month = quarter_months[triwulan]
            start_date = datetime(tahun, start_month, 1).date()
            if end_month == 12:
                end_date = datetime(tahun + 1, 1, 1).date() - timedelta(days=1)
            else:
                end_date = datetime(tahun, end_month + 1, 1).date() - timedelta(days=1)
            periode_label = f'Triwulan {triwulan} {tahun}'
        except (ValueError, TypeError):
            return HttpResponse('Invalid quarter', status=400)
    
    elif periode_type == 'semester':
        semester_months = {
            1: (1, 6),
            2: (7, 12),
        }
        try:
            semester = int(periode)
            if semester not in semester_months:
                return HttpResponse('Invalid semester', status=400)
            start_month, end_month = semester_months[semester]
            start_date = datetime(tahun, start_month, 1).date()
            if end_month == 12:
                end_date = datetime(tahun + 1, 1, 1).date() - timedelta(days=1)
            else:
                end_date = datetime(tahun, end_month + 1, 1).date() - timedelta(days=1)
            periode_label = f'Semester {semester} {tahun}'
        except (ValueError, TypeError):
            return HttpResponse('Invalid semester', status=400)
    
    elif periode_type == 'tahunan':
        start_date = datetime(tahun, 1, 1).date()
        end_date = datetime(tahun, 12, 31).date()
        periode_label = f'{tahun}'
    
    else:
        return HttpResponse('Invalid periode type', status=400)
    
    
    # Query tikets
    tikets = Tiket.objects.filter(
        tgl_transfer__isnull=False,
        tgl_transfer__date__gte=start_date,
        tgl_transfer__date__lte=end_date
    ).select_related(
        'id_periode_data__id_sub_jenis_data_ilap__id_ilap',
        'id_periode_data__id_sub_jenis_data_ilap__id_jenis_tabel'
    ).order_by('-tgl_transfer')
    
    # Use django-import-export to generate XLSX
    resource = TiketExportResource()
    dataset = resource.export(tikets)
    
    # Create Excel workbook using openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Tikets"
    
    # Write headers
    headers = dataset.headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Write data rows
    for row_idx, row in enumerate(dataset, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_data = excel_file.getvalue()
    
    # Create HTTP response
    response = HttpResponse(
        excel_data,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Laporan_Pengendalian_Mutu_{periode_label.replace(" ", "_")}.xlsx"'
    return response
