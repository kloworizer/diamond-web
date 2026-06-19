"""Laporan SLA Identifikasi view - Matching SLA Report."""

from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.decorators.http import require_GET, require_http_methods
from django.views.generic import TemplateView
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_protect
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook

from ..models.tiket import Tiket
from ..forms.laporan_sla_identifikasi import LaporanSLAIdentifikasiFilterForm, LaporanSLAIdentifikasiExportResource


def _is_pide_user(user):
    """Check if user is PIDE user or admin."""
    return user.is_superuser or user.is_staff or user.groups.filter(name__in=['user_pide', 'admin', 'admin_pide']).exists()


def _get_filtered_tikets(params):
    """Utility function to filter tikets based on request parameters.

    Applies filters for date range (tgl_mulai, tgl_akhir filtered on
    tgl_rekam_pide), ILAP, Jenis Data, Sub Jenis Data, and Tabel I.
    Results are ordered by tgl_kirim_pide ascending.

    Args:
        params: QueryDict or dict-like object containing filter parameters.
            - tgl_mulai (str, optional): Start datetime in YYYY-MM-DDTHH:MM format.
            - tgl_akhir (str, optional): End datetime in YYYY-MM-DDTHH:MM format.
            - id_ilap (str, optional): ILAP ID to filter by.
            - id_jenis_data (str, optional): Jenis Data ID to filter by.
            - nama_sub_jenis_data (str, optional): Sub-jenis data name to filter by.
            - nama_tabel_I (str, optional): Tabel I name to filter by.

    Returns:
        QuerySet: Filtered and ordered Tiket QuerySet with select_related
            on id_periode_data__id_sub_jenis_data_ilap__id_ilap and
            id_periode_data__id_sub_jenis_data_ilap__id_jenis_tabel.
    """
    tgl_mulai_str = params.get('tgl_mulai')
    tgl_akhir_str = params.get('tgl_akhir')
    id_ilap = params.get('id_ilap')
    id_jenis_data = params.get('id_jenis_data')
    nama_sub_jenis_data = params.get('nama_sub_jenis_data')
    nama_tabel_I = params.get('nama_tabel_I')

    # Query base
    tikets = Tiket.objects.all().select_related(
        'id_periode_data__id_sub_jenis_data_ilap__id_ilap',
        'id_periode_data__id_sub_jenis_data_ilap__id_jenis_tabel'
    )
    
    # Filter by date range
    if tgl_mulai_str:
        try:
            # datetime-local format: YYYY-MM-DDTHH:MM
            tgl_mulai = datetime.strptime(tgl_mulai_str, '%Y-%m-%dT%H:%M')
            tikets = tikets.filter(tgl_rekam_pide__gte=tgl_mulai)
        except ValueError:
            pass
            
    if tgl_akhir_str:
        try:
            tgl_akhir = datetime.strptime(tgl_akhir_str, '%Y-%m-%dT%H:%M')
            tikets = tikets.filter(tgl_rekam_pide__lte=tgl_akhir)
        except ValueError:
            pass
            
    # Filter by ILAP
    if id_ilap and id_ilap != 'all' and id_ilap != '':
        tikets = tikets.filter(id_periode_data__id_sub_jenis_data_ilap__id_ilap_id=id_ilap)
        
    # Filter by Jenis Data
    if id_jenis_data and id_jenis_data != 'all' and id_jenis_data != '':
        tikets = tikets.filter(id_periode_data__id_sub_jenis_data_ilap_id=id_jenis_data)
        
    # Filter by Subjenis Data
    if nama_sub_jenis_data and nama_sub_jenis_data != 'all' and nama_sub_jenis_data != '':
        tikets = tikets.filter(id_periode_data__id_sub_jenis_data_ilap__nama_sub_jenis_data=nama_sub_jenis_data)
        
    # Filter by Tabel I
    if nama_tabel_I and nama_tabel_I != 'all' and nama_tabel_I != '':
        tikets = tikets.filter(id_periode_data__id_sub_jenis_data_ilap__nama_tabel_I=nama_tabel_I)

    # Order by sent to PIDE date (chronological)
    return tikets.order_by('tgl_kirim_pide')


class LaporanSLAIdentifikasiView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Display Laporan SLA Identifikasi by filtering tikets
    based on a date range and other parameters.
    """
    template_name = 'laporan_sla_identifikasi/list.html'
    
    def test_func(self):
        """Check if the current user is a PIDE user or admin.

        Returns:
            bool: True if the user has PIDE or admin privileges.
        """
        return _is_pide_user(self.request.user)
    
    def get_context_data(self, **kwargs):
        """Add the SLA identifikasi filter form to the template context.

        Args:
            **kwargs: Additional context arguments passed to the parent.

        Returns:
            dict: Context dict with an initialized 'form' instance.
        """
        context = super().get_context_data(**kwargs)
        context['form'] = LaporanSLAIdentifikasiFilterForm()
        return context


@login_required
@user_passes_test(_is_pide_user)
@require_http_methods(["GET", "POST"])
@csrf_protect
def laporan_sla_identifikasi_data(request):
    """DataTables server-side endpoint for SLA Identifikasi Report.
    
    Filters tikets by:
    - tgl_rekam_pide within specified tgl_mulai and tgl_akhir
    - id_ilap, id_jenis_data, nama_sub_jenis_data, nama_tabel_I
    
    Returns JSON with tiket data.
    """
    params = request.POST if request.method == 'POST' else request.GET
    
    try:
        draw = int(params.get('draw', 1))
        start = int(params.get('start', 0))
        length = int(params.get('length', 10))
    except (ValueError, TypeError):
        draw, start, length = 1, 0, 10
    
    # Get filtered tikets using helper
    tikets = _get_filtered_tikets(params)
    
    records_total = Tiket.objects.count()
    records_filtered = tikets.count()
    
    # Pagination
    tikets_paginated = tikets[start:start + length]
    
    # Build response data
    data = []
    for tiket in tikets_paginated:
        sub_jenis_data = tiket.id_periode_data.id_sub_jenis_data_ilap if tiket.id_periode_data else None
        if not sub_jenis_data:
            continue
            
        ilap = sub_jenis_data.id_ilap
        
        # Logic for calculated fields (matching LaporanSLAIdentifikasiExportResource)
        if tiket.tgl_rekam_pide and tiket.tgl_transfer:
            diff = (tiket.tgl_transfer - tiket.tgl_rekam_pide).days + 1
            sla_identifikasi = f"{diff} hari"
        else:
            sla_identifikasi = ""
        if tiket.tgl_rekam_pide:
            tgl_mulai_identifikasi = tiket.tgl_rekam_pide.strftime('%d/%m/%Y')
        else:
            tgl_mulai_identifikasi = ""
        if tiket.tgl_transfer:
            tgl_transfer = tiket.tgl_transfer.strftime('%d/%m/%Y')
        else:
            tgl_transfer = ""

        row = {
            'nama_ilap': ilap.nama_ilap if ilap else '',
            'nama_jenis_data': sub_jenis_data.nama_jenis_data,
            'nama_sub_jenis_data': sub_jenis_data.nama_sub_jenis_data,
            'nama_tabel_I': sub_jenis_data.nama_tabel_I,
            'nomor_tiket': tiket.nomor_tiket,
            'tgl_mulai_identifikasi': tgl_mulai_identifikasi,
            'tgl_transfer': tgl_transfer,
            'sla_identifikasi': sla_identifikasi,
        }
        data.append(row)
    
    return JsonResponse({
        'draw': draw,
        'recordsTotal': records_total,
        'recordsFiltered': records_filtered,
        'data': data
    })


@login_required
@user_passes_test(_is_pide_user)
@require_GET
@csrf_protect
def laporan_sla_identifikasi_export(request):
    """Export Laporan SLA Identifikasi to XLSX file.

    Uses the LaporanSLAIdentifikasiExportResource to generate an XLSX file from
    tikets filtered by the provided GET parameters (date range, ILAP,
    Jenis Data, Sub Jenis Data, Tabel I).

    Args:
        request: The HTTP GET request object. Parameters are passed to
            _get_filtered_tikets() for filtering.
            - tgl_mulai (str, optional): Start date in YYYY-MM-DDTHH:MM format.
            - tgl_akhir (str, optional): End date in YYYY-MM-DDTHH:MM format.
            - id_ilap (str, optional): ILAP ID filter.
            - id_jenis_data (str, optional): Jenis Data ID filter.
            - nama_sub_jenis_data (str, optional): Sub-jenis data name filter.
            - nama_tabel_I (str, optional): Tabel I name filter.

    Returns:
        HttpResponse: An XLSX file download response with the exported data.
    """
    # Get filtered tikets using helper
    tikets = _get_filtered_tikets(request.GET)
    
    # Use LaporanSLAIdentifikasiExportResource
    resource = LaporanSLAIdentifikasiExportResource()
    dataset = resource.export(tikets)
    
    # Create Excel workbook using openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan SLA Identifikasi"
    
    # Write headers
    headers = dataset.headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Write data rows
    for row_idx, row in enumerate(dataset, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_data = excel_file.getvalue()
    
    # Create filename
    filename = "Laporan_SLA_Identifikasi"
    tgl_mulai_str = request.GET.get('tgl_mulai')
    tgl_akhir_str = request.GET.get('tgl_akhir')
    if tgl_mulai_str and tgl_akhir_str:
        filename += f"_{tgl_mulai_str[:10]}_ke_{tgl_akhir_str[:10]}"
    
    # Create HTTP response
    response = HttpResponse(
        excel_data,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response

